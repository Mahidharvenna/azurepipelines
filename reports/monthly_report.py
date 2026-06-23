#!/usr/bin/env python3
"""
Monthly Guidewire login report -> Excel -> email.

Counts lines containing "User Login" in the PC (and optionally BC/CC/CM)
logs stored in Loki, per environment, for the last full calendar month,
then emails an Excel workbook.

Mirrors a typical Grafana "user logins" panel: same label selectors,
same `|= "User Login"` line match -- so the report reconciles with the
dashboard.

Driven entirely by environment variables (set by the pipeline via a
Variable Group). No secrets are hard-coded.
"""

import os
import smtplib
from datetime import datetime, timedelta, timezone
from email.message import EmailMessage
from email.utils import formatdate
from io import BytesIO

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference

# ---------------------------------------------------------------------------
# CONFIG (all overridable via env vars from the pipeline's Variable Group)
# ---------------------------------------------------------------------------
LOKI_URL = os.environ.get("LOKI_URL", "https://your-loki-host.example.com:3100")
# TLS: if Loki is internal HTTPS and the build agent doesn't trust the cert,
# set LOKI_VERIFY_TLS=false OR point LOKI_CA_BUNDLE at your internal CA .pem.
_verify_env = os.environ.get("LOKI_VERIFY_TLS", "true").strip().lower()
LOKI_CA_BUNDLE = os.environ.get("LOKI_CA_BUNDLE", "").strip()
if LOKI_CA_BUNDLE:
    VERIFY = LOKI_CA_BUNDLE
else:
    VERIFY = _verify_env not in ("false", "0", "no")

SMTP_HOST = os.environ["SMTP_HOST"]
SMTP_PORT = int(os.environ.get("SMTP_PORT", "25"))
SMTP_USER = os.environ.get("SMTP_USER", "")
SMTP_PASS = os.environ.get("SMTP_PASS", "")
SMTP_TLS = os.environ.get("SMTP_TLS", "false").strip().lower() in ("true", "1", "yes")

FROM_ADDR = os.environ["FROM_ADDR"]
TO_ADDRS = [a.strip() for a in os.environ["TO_ADDRS"].split(",") if a.strip()]

# Envs use the SAME casing as your Loki `env` label (e.g. DEV1, QA7, UAT1).
ENVS = [e.strip() for e in os.environ.get("ENVS", "DEV1,QA1,UAT1,PROD1").split(",") if e.strip()]

# Products to report. Each maps to a Loki `job` label + filename fragment.
# Default is PC only.
PRODUCTS = [p.strip().lower() for p in os.environ.get("PRODUCTS", "pc").split(",") if p.strip()]

# The Loki `project` label value that scopes all your Guidewire logs.
PROJECT = os.environ.get("LOKI_PROJECT", "myproject")

OUTPUT_DIR = os.environ.get("OUTPUT_DIR", ".")

# job label + filename fragment per product. Adjust to your own label scheme;
# confirm the job names via the Grafana Label browser.
PRODUCT_META = {
    "pc": {"job": "pclogs", "frag": "pc", "label": "PolicyCenter"},
    "bc": {"job": "bclogs", "frag": "bc", "label": "BillingCenter"},
    "cc": {"job": "cclogs", "frag": "cc", "label": "ClaimCenter"},
    "cm": {"job": "cmlogs", "frag": "cm", "label": "ContactManager"},
}

# ---------------------------------------------------------------------------
# DATE WINDOW: last full calendar month, in UTC
# ---------------------------------------------------------------------------
# Optional override for testing: TEST_MONTH=YYYY-MM
_override = os.environ.get("TEST_MONTH", "").strip()
if _override:
    start = datetime.strptime(_override + "-01", "%Y-%m-%d").replace(tzinfo=timezone.utc)
    end = (start + timedelta(days=32)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
else:
    now = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    end = now.replace(day=1)                          # 00:00 on the 1st of THIS month
    start = (end - timedelta(days=1)).replace(day=1)  # 00:00 on the 1st of LAST month

month_label = start.strftime("%B %Y")                 # e.g. "May 2026"
gen_stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

# ---------------------------------------------------------------------------
# LOKI QUERY  (match: |= `User Login`)
# ---------------------------------------------------------------------------
QUERY_TMPL = (
    'sum(count_over_time('
    '{{project="{project}", job="{job}", env="{env}", filename=~".*{frag}.log"}} '
    '|= `User Login` [1d]))'
)


def fetch_daily(env, product):
    """Return list[(date, count)] for one env+product over the report window."""
    meta = PRODUCT_META[product]
    query = QUERY_TMPL.format(project=PROJECT, job=meta["job"], env=env, frag=meta["frag"])
    r = requests.get(
        f"{LOKI_URL}/loki/api/v1/query_range",
        params={
            "query": query,
            "start": int(start.timestamp() * 1e9),
            "end": int(end.timestamp() * 1e9),
            "step": "1d",
        },
        timeout=120,
        verify=VERIFY,
    )
    r.raise_for_status()
    result = r.json()["data"]["result"]
    if not result:
        return []
    out = []
    for ts, val in result[0]["values"]:
        d = datetime.fromtimestamp(float(ts), tz=timezone.utc).date()
        out.append((d, int(float(val))))
    return out


# Collect: data[(env, product)] = {"daily": [(date,count)], "total": int}
data = {}
for env in ENVS:
    for product in PRODUCTS:
        daily = fetch_daily(env, product)
        data[(env, product)] = {
            "daily": daily,
            "total": sum(c for _, c in daily),
        }

# ---------------------------------------------------------------------------
# BUILD WORKBOOK
# ---------------------------------------------------------------------------
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="305496")
title_font = Font(bold=True, size=14)
bold = Font(bold=True)
thin = Border(*[Side(border_style="thin", color="CCCCCC")] * 4)

wb = Workbook()

# --- Sheet 1: Summary ------------------------------------------------------
ws = wb.active
ws.title = "Summary"
ws["A1"] = "Guidewire Login Report"
ws["A1"].font = title_font
ws.merge_cells("A1:C1")
ws["A2"] = f"Period: {start:%Y-%m-%d} through {(end - timedelta(days=1)):%Y-%m-%d}  ({month_label})"
ws["A3"] = f"Source: Loki {LOKI_URL}  |  Match: |= \"User Login\"  |  Generated: {gen_stamp}"

heads = ["Environment", "Product", "Total Logins"]
for col, h in enumerate(heads, start=1):
    c = ws.cell(row=5, column=col, value=h)
    c.font, c.fill, c.border = header_font, header_fill, thin
    c.alignment = Alignment(horizontal="center")

row = 6
first_data_row = row
for env in ENVS:
    for product in PRODUCTS:
        ws.cell(row=row, column=1, value=env).border = thin
        ws.cell(row=row, column=2, value=PRODUCT_META[product]["label"]).border = thin
        tc = ws.cell(row=row, column=3, value=data[(env, product)]["total"])
        tc.number_format = "#,##0"
        tc.border = thin
        row += 1
last_data_row = row - 1

ws.cell(row=row, column=1, value="TOTAL").font = bold
tc = ws.cell(row=row, column=3, value=f"=SUM(C{first_data_row}:C{last_data_row})")
tc.number_format = "#,##0"
tc.font = bold

ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 16
ws.freeze_panes = "A6"

# --- Sheet 2: Daily --------------------------------------------------------
ws2 = wb.create_sheet("Daily")
ws2["A1"] = "Daily Login Counts"
ws2["A1"].font = title_font

combos = [(env, product) for env in ENVS for product in PRODUCTS]
ws2.cell(row=3, column=1, value="Date").font = header_font
ws2.cell(row=3, column=1).fill = header_fill
for col, (env, product) in enumerate(combos, start=2):
    name = env if len(PRODUCTS) == 1 else f"{env}/{product.upper()}"
    c = ws2.cell(row=3, column=col, value=name)
    c.font, c.fill = header_font, header_fill
    c.alignment = Alignment(horizontal="center")

all_dates = sorted({d for v in data.values() for d, _ in v["daily"]})
for i, day in enumerate(all_dates, start=4):
    ws2.cell(row=i, column=1, value=day).number_format = "yyyy-mm-dd"
    for col, (env, product) in enumerate(combos, start=2):
        match = next((c for d, c in data[(env, product)]["daily"] if d == day), 0)
        cell = ws2.cell(row=i, column=col, value=match)
        cell.number_format = "#,##0"

ws2.column_dimensions["A"].width = 13
for col in range(2, len(combos) + 2):
    ws2.column_dimensions[ws2.cell(row=3, column=col).column_letter].width = 14
ws2.freeze_panes = "B4"

if all_dates:
    chart = LineChart()
    chart.title = f"Daily Logins - {month_label}"
    chart.y_axis.title = "Logins"
    chart.x_axis.title = "Date"
    chart.height, chart.width = 9, 20
    vals = Reference(ws2, min_col=2, max_col=len(combos) + 1,
                     min_row=3, max_row=3 + len(all_dates))
    cats = Reference(ws2, min_col=1, max_col=1,
                     min_row=4, max_row=3 + len(all_dates))
    chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)
    anchor = ws2.cell(row=3, column=len(combos) + 3).column_letter + "3"
    ws2.add_chart(chart, anchor)

# ---------------------------------------------------------------------------
# SAVE
# ---------------------------------------------------------------------------
os.makedirs(OUTPUT_DIR, exist_ok=True)
filename = f"gw-logins-{start:%Y-%m}.xlsx"
xlsx_path = os.path.join(OUTPUT_DIR, filename)
buf = BytesIO()
wb.save(buf)
xlsx_bytes = buf.getvalue()
with open(xlsx_path, "wb") as f:
    f.write(xlsx_bytes)
print(f"Wrote {xlsx_path} ({len(xlsx_bytes):,} bytes)")

# ---------------------------------------------------------------------------
# EMAIL
# ---------------------------------------------------------------------------
rows_html = "".join(
    f"<li><b>{env} {PRODUCT_META[product]['label']}</b>: "
    f"{data[(env, product)]['total']:,} logins</li>"
    for env in ENVS for product in PRODUCTS
)
html = f"""\
<html><body style="font-family:sans-serif">
<p>Hi team,</p>
<p>Attached is the Guidewire login report for <b>{month_label}</b>.</p>
<ul>{rows_html}</ul>
<p style="color:#888;font-size:0.85em">
Source: Loki at {LOKI_URL}. Match: <code>|= "User Login"</code>.
Generated {gen_stamp} (automated).
</p>
</body></html>
"""

msg = EmailMessage()
msg["Subject"] = f"Guidewire Login Report - {month_label}"
msg["From"] = FROM_ADDR
msg["To"] = ", ".join(TO_ADDRS)
msg["Date"] = formatdate(localtime=True)
msg.set_content("This report requires an HTML-capable mail client. See the attached Excel file.")
msg.add_alternative(html, subtype="html")
msg.add_attachment(
    xlsx_bytes,
    maintype="application",
    subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    filename=filename,
)

with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
    if SMTP_TLS:
        s.starttls()
    if SMTP_USER:
        s.login(SMTP_USER, SMTP_PASS)
    s.send_message(msg)

print(f"Sent '{filename}' to {len(TO_ADDRS)} recipient(s): {', '.join(TO_ADDRS)}")
